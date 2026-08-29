"""
plate_detector.py — Standalone License Plate Detection Module
================================================================
Opens a camera feed, detects number plates with a YOLO model,
reads the plate text with EasyOCR, draws a live tracking box +
plate number on screen, and logs every new plate read to an
Excel file (plate_log.xlsx).

This module has NO database dependency — it's a self-contained
detector you can drop into any project.

Requirements:
    pip install opencv-python ultralytics easyocr openpyxl numpy

Usage:
    python plate_detector.py            # use default webcam (index 0)
    python plate_detector.py 1          # use camera index 1
    python plate_detector.py photo.jpg  # run on a single image instead of a camera
"""

import os
import re
import sys
import time
from collections import defaultdict
from datetime import datetime

import cv2
import numpy as np
from ultralytics import YOLO
import easyocr
from openpyxl import Workbook, load_workbook

# ── CONFIG ──────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# Path to the trained YOLO plate-detection weights.
# Defaults to the "best.pt" that ships with this project (runs/detect/train/weights/best.pt).
# Point this at your own weights if you have them elsewhere.
MODEL_PATH = os.path.join(SCRIPT_DIR, "runs", "detect", "train", "weights", "best.pt")

EXCEL_PATH = os.path.join(SCRIPT_DIR, "plate_log.xlsx")

DET_CONFIDENCE   = 0.5    # YOLO detection confidence threshold
MIN_OCR_CONFIDENCE = 0.4  # minimum OCR confidence to accept/log a read
DEBOUNCE_SECONDS = 5.0    # don't re-log the same plate within this window
WINDOW_NAME      = "Plate Detector"
# ────────────────────────────────────────────────────────────────

print("Loading YOLO model...")
if not os.path.exists(MODEL_PATH):
    print(f"ERROR: model weights not found at: {MODEL_PATH}")
    print("Update MODEL_PATH at the top of this script to point at your best.pt")
    sys.exit(1)
model = YOLO(MODEL_PATH)

print("Loading EasyOCR (first run may take a moment)...")
reader = easyocr.Reader(['en'], gpu=False, verbose=False)
print("Ready.\n")


# ═══════════════════════════════════════════════════════════════
#  IMAGE / OCR HELPERS
# ═══════════════════════════════════════════════════════════════

def upscale(img: np.ndarray, min_h: int = 100) -> np.ndarray:
    h, w = img.shape[:2]
    if h < min_h:
        scale = min_h / h
        img = cv2.resize(img, (int(w * scale), int(h * scale)),
                          interpolation=cv2.INTER_LANCZOS4)
    return img


def preprocess_plate(crop: np.ndarray) -> np.ndarray:
    """Upscale + contrast/sharpen boost so OCR reads plates more reliably."""
    crop = upscale(crop, min_h=100)
    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(4, 4))
    gray = clahe.apply(gray)
    kernel = np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]], dtype=np.float32)
    gray = cv2.filter2D(gray, -1, kernel)
    return cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)


def clean_text(text: str) -> str:
    return re.sub(r"[^A-Z0-9\-]", "", text.upper().strip())


def ocr_plate(crop: np.ndarray):
    """Returns (plate_text, avg_confidence)."""
    if crop.size == 0:
        return "", 0.0
    crop = preprocess_plate(crop)
    results = reader.readtext(
        crop,
        allowlist="ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-",
        detail=1,
    )
    hits = [(r[0][0][0], clean_text(r[1]), r[2]) for r in results if clean_text(r[1])]
    hits.sort(key=lambda h: h[0])  # left → right
    text = "".join(h[1] for h in hits)
    conf = (sum(h[2] for h in hits) / len(hits)) if hits else 0.0
    return text, conf


def draw_box(frame, x1, y1, x2, y2, plate_text, det_conf, ocr_conf, is_new):
    color = (0, 220, 0) if plate_text else (0, 165, 255)
    if is_new:
        color = (0, 255, 255) 
    cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)

    label = plate_text if plate_text else "reading..."
    display = f"{label}  [det:{det_conf:.2f} ocr:{ocr_conf:.2f}]"

    (tw, th), _ = cv2.getTextSize(display, cv2.FONT_HERSHEY_SIMPLEX, 0.65, 2)
    bg_y1 = max(y1 - th - 12, 0)
    cv2.rectangle(frame, (x1, bg_y1), (x1 + tw + 8, y1), color, -1)
    cv2.putText(frame, display, (x1 + 4, y1 - 5),
                cv2.FONT_HERSHEY_SIMPLEX, 0.65, (0, 0, 0), 2)


# ═══════════════════════════════════════════════════════════════
#  EXCEL LOGGING
# ═══════════════════════════════════════════════════════════════

def init_excel():
    """Create the log workbook with headers if it doesn't exist yet."""
    if os.path.exists(EXCEL_PATH):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Plate Log"
    ws.append(["Date", "Time", "Plate Number", "Detection Confidence", "OCR Confidence"])
    wb.save(EXCEL_PATH)


def log_plate(plate: str, det_conf: float, ocr_conf: float):
    """Append one row to the Excel log."""
    init_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    now = datetime.now()
    ws.append([
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
        plate,
        round(det_conf, 2),
        round(ocr_conf, 2),
    ])
    wb.save(EXCEL_PATH)
    print(f"[{now:%H:%M:%S}] Logged plate: {plate}  (det={det_conf:.2f}, ocr={ocr_conf:.2f})")



def run_on_camera(camera_index: int = 0):
    init_excel()
    cap = cv2.VideoCapture(camera_index)
    if not cap.isOpened():
        print(f"ERROR: cannot open camera index {camera_index}")
        return

    last_seen = defaultdict(float)  # plate text -> last time it was logged
    print(f"Camera {camera_index} opened. Press 'q' to quit.\n")

    while True:
        ret, frame = cap.read()
        if not ret:
            print("Failed to read frame from camera.")
            break

        results = model(frame, conf=DET_CONFIDENCE, verbose=False)
        boxes = results[0].boxes

        if boxes is not None and len(boxes) > 0:
            for box in boxes:
                x1, y1, x2, y2 = map(int, box.xyxy[0])
                det_conf = float(box.conf[0])
                pad = 6
                crop = frame[max(0, y1 - pad):y2 + pad, max(0, x1 - pad):x2 + pad]

                plate_text, ocr_conf = ocr_plate(crop)
                is_new = False

                if plate_text and ocr_conf >= MIN_OCR_CONFIDENCE:
                    now = time.time()
                    if now - last_seen[plate_text] > DEBOUNCE_SECONDS:
                        last_seen[plate_text] = now
                        log_plate(plate_text, det_conf, ocr_conf)
                        is_new = True

                draw_box(frame, x1, y1, x2, y2, plate_text, det_conf, ocr_conf, is_new)

        cv2.putText(frame, "Press 'q' to quit", (10, 25),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)
        cv2.imshow(WINDOW_NAME, frame)

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()
    print(f"\nDone. Plate log saved to: {EXCEL_PATH}")


def run_on_image(image_path: str):
    if not os.path.exists(image_path):
        print(f"ERROR: file not found — {image_path}")
        return

    frame = cv2.imread(image_path)
    if frame is None:
        print(f"ERROR: could not read image — {image_path}")
        return

    results = model(frame, conf=DET_CONFIDENCE, verbose=False)
    boxes = results[0].boxes

    if boxes is None or len(boxes) == 0:
        print("No plates detected.")
    else:
        for box in boxes:
            x1, y1, x2, y2 = map(int, box.xyxy[0])
            det_conf = float(box.conf[0])
            pad = 6
            crop = frame[max(0, y1 - pad):y2 + pad, max(0, x1 - pad):x2 + pad]
            plate_text, ocr_conf = ocr_plate(crop)

            is_new = False
            if plate_text and ocr_conf >= MIN_OCR_CONFIDENCE:
                log_plate(plate_text, det_conf, ocr_conf)
                is_new = True

            draw_box(frame, x1, y1, x2, y2, plate_text, det_conf, ocr_conf, is_new)

    cv2.imshow(WINDOW_NAME, frame)
    print("Press any key to close...")
    cv2.waitKey(0)
    cv2.destroyAllWindows()
    print(f"\nDone. Plate log saved to: {EXCEL_PATH}")


# ── ENTRY POINT ──────────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) > 1:
        arg = sys.argv[1]
        if arg.isdigit():
            run_on_camera(int(arg))
        else:
            run_on_image(arg)
    else:
        run_on_camera(0)