"""
app.py — License Plate Detector (Streamlit, image input)
==========================================================
Browser-based version of plate_detector.py. Upload a photo,
the app detects plates with YOLO, reads the text with EasyOCR,
draws boxes on the image, and logs new reads to plate_log.xlsx.

Requirements:
    pip install streamlit opencv-python-headless ultralytics easyocr openpyxl numpy pillow pandas

Run:
    streamlit run app.py
"""

import os
import re
import time
from datetime import datetime

import cv2
import numpy as np
import pandas as pd
import streamlit as st
from PIL import Image
from openpyxl import Workbook, load_workbook

# ── CONFIG ──────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_MODEL_PATH = os.path.join(SCRIPT_DIR, "runs", "detect", "train", "weights", "best.pt")
EXCEL_PATH = os.path.join(SCRIPT_DIR, "plate_log.xlsx")
DEBOUNCE_SECONDS = 5.0  # don't re-log the same plate within this window
# ────────────────────────────────────────────────────────────────


# ═══════════════════════════════════════════════════════════════
#  CACHED MODEL LOADERS
# ═══════════════════════════════════════════════════════════════

@st.cache_resource(show_spinner="Loading YOLO model...")
def load_yolo(model_path: str):
    from ultralytics import YOLO
    if not os.path.exists(model_path):
        return None
    return YOLO(model_path)


@st.cache_resource(show_spinner="Loading EasyOCR (first run may take a moment)...")
def load_ocr_reader():
    import easyocr
    return easyocr.Reader(["en"], gpu=False, verbose=False)


# ═══════════════════════════════════════════════════════════════
#  IMAGE / OCR HELPERS  (unchanged logic from plate_detector.py)
# ═══════════════════════════════════════════════════════════════

def upscale(img: np.ndarray, min_h: int = 100) -> np.ndarray:
    h, w = img.shape[:2]
    if h < min_h:
        scale = min_h / h
        img = cv2.resize(img, (int(w * scale), int(h * scale)),
                          interpolation=cv2.INTER_LANCZOS4)
    return img


def preprocess_plate(crop: np.ndarray) -> np.ndarray:
    """Upscale + contrast/sharpen boost so OCR reads plates more reliably."""
    crop = upscale(crop, min_h=100)
    gray = cv2.cvtColor(crop, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(4, 4))
    gray = clahe.apply(gray)
    kernel = np.array([[0, -1, 0], [-1, 5, -1], [0, -1, 0]], dtype=np.float32)
    gray = cv2.filter2D(gray, -1, kernel)
    return cv2.cvtColor(gray, cv2.COLOR_GRAY2BGR)


def clean_text(text: str) -> str:
    return re.sub(r"[^A-Z0-9\-]", "", text.upper().strip())


def ocr_plate(reader, crop: np.ndarray):
    """Returns (plate_text, avg_confidence)."""
    if crop.size == 0:
        return "", 0.0
    crop = preprocess_plate(crop)
    results = reader.readtext(
        crop,
        allowlist="ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-",
        detail=1,
    )
    hits = [(r[0][0][0], clean_text(r[1]), r[2]) for r in results if clean_text(r[1])]
    hits.sort(key=lambda h: h[0])  # left → right
    text = "".join(h[1] for h in hits)
    conf = (sum(h[2] for h in hits) / len(hits)) if hits else 0.0
    return text, conf


def draw_box(frame, x1, y1, x2, y2, plate_text, det_conf, ocr_conf, is_new):
    color = (0, 220, 0) if plate_text else (0, 165, 255)
    if is_new:
        color = (0, 255, 255)  # highlight freshly logged plates
    cv2.rectangle(frame, (x1, y1), (x2, y2), color, 2)

    label = plate_text if plate_text else "reading..."
    display = f"{label}  [det:{det_conf:.2f} ocr:{ocr_conf:.2f}]"

    (tw, th), _ = cv2.getTextSize(display, cv2.FONT_HERSHEY_SIMPLEX, 0.65, 2)
    bg_y1 = max(y1 - th - 12, 0)
    cv2.rectangle(frame, (x1, bg_y1), (x1 + tw + 8, y1), color, -1)
    cv2.putText(frame, display, (x1 + 4, y1 - 5),
                cv2.FONT_HERSHEY_SIMPLEX, 0.65, (0, 0, 0), 2)


# ═══════════════════════════════════════════════════════════════
#  EXCEL LOGGING  (unchanged logic from plate_detector.py)
# ═══════════════════════════════════════════════════════════════

def init_excel():
    if os.path.exists(EXCEL_PATH):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Plate Log"
    ws.append(["Date", "Time", "Plate Number", "Detection Confidence", "OCR Confidence"])
    wb.save(EXCEL_PATH)


def log_plate(plate: str, det_conf: float, ocr_conf: float):
    init_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    now = datetime.now()
    ws.append([
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
        plate,
        round(det_conf, 2),
        round(ocr_conf, 2),
    ])
    wb.save(EXCEL_PATH)


def read_log_df():
    if not os.path.exists(EXCEL_PATH):
        return pd.DataFrame(columns=["Date", "Time", "Plate Number", "Detection Confidence", "OCR Confidence"])
    return pd.read_excel(EXCEL_PATH)


# ═══════════════════════════════════════════════════════════════
#  CORE DETECTION (single image)
# ═══════════════════════════════════════════════════════════════

def detect_plates(frame_bgr, model, reader, det_confidence, min_ocr_confidence, last_seen):
    """Runs YOLO + OCR on one BGR frame, draws boxes, logs new plates.
    Returns (annotated_frame, list_of_detections)."""
    results = model(frame_bgr, conf=det_confidence, verbose=False)
    boxes = results[0].boxes
    detections = []

    if boxes is not None and len(boxes) > 0:
        for box in boxes:
            x1, y1, x2, y2 = map(int, box.xyxy[0])
            det_conf = float(box.conf[0])
            pad = 6
            crop = frame_bgr[max(0, y1 - pad):y2 + pad, max(0, x1 - pad):x2 + pad]

            plate_text, ocr_conf = ocr_plate(reader, crop)
            is_new = False

            if plate_text and ocr_conf >= min_ocr_confidence:
                now = time.time()
                if now - last_seen.get(plate_text, 0) > DEBOUNCE_SECONDS:
                    last_seen[plate_text] = now
                    log_plate(plate_text, det_conf, ocr_conf)
                    is_new = True

            draw_box(frame_bgr, x1, y1, x2, y2, plate_text, det_conf, ocr_conf, is_new)
            detections.append({
                "Plate Number": plate_text or "(unreadable)",
                "Detection Confidence": round(det_conf, 2),
                "OCR Confidence": round(ocr_conf, 2),
                "Logged": "Yes" if is_new else "No",
            })

    return frame_bgr, detections


# ═══════════════════════════════════════════════════════════════
#  STREAMLIT APP
# ═══════════════════════════════════════════════════════════════

st.set_page_config(page_title="License Plate Detector", page_icon="🚘", layout="wide")
st.title("🚘 License Plate Detector")
st.caption("Upload a photo to detect and read license plates. Every new plate read is logged to an Excel file.")

if "last_seen" not in st.session_state:
    st.session_state.last_seen = {}

with st.sidebar:
    st.header("Settings")
    model_path = st.text_input("YOLO weights path", value=DEFAULT_MODEL_PATH)
    det_confidence = st.slider("Detection confidence threshold", 0.0, 1.0, 0.5, 0.05)
    min_ocr_confidence = st.slider("Minimum OCR confidence to log", 0.0, 1.0, 0.4, 0.05)
    st.divider()
    if st.button("Clear debounce memory"):
        st.session_state.last_seen = {}
        st.success("Cleared — the next matching plate will be logged again.")

model = load_yolo(model_path)
if model is None:
    st.error(f"Model weights not found at: `{model_path}`\n\nUpdate the path in the sidebar to point at your `best.pt`.")
    st.stop()

reader = load_ocr_reader()

uploaded_file = st.file_uploader("Upload an image", type=["jpg", "jpeg", "png", "bmp", "webp"])

if uploaded_file is not None:
    pil_img = Image.open(uploaded_file).convert("RGB")
    frame_bgr = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)

    with st.spinner("Detecting plates..."):
        annotated_bgr, detections = detect_plates(
            frame_bgr, model, reader, det_confidence, min_ocr_confidence, st.session_state.last_seen
        )

    annotated_rgb = cv2.cvtColor(annotated_bgr, cv2.COLOR_BGR2RGB)

    col1, col2 = st.columns([2, 1])
    with col1:
        st.image(annotated_rgb, caption="Detection result", use_container_width=True)
    with col2:
        st.subheader("Detections")
        if detections:
            st.dataframe(pd.DataFrame(detections), use_container_width=True, hide_index=True)
        else:
            st.info("No plates detected in this image.")

st.divider()
st.subheader("📄 Plate Log")
log_df = read_log_df()
st.dataframe(log_df, use_container_width=True, hide_index=True)

if os.path.exists(EXCEL_PATH):
    with open(EXCEL_PATH, "rb") as f:
        st.download_button(
            "Download plate_log.xlsx",
            data=f.read(),
            file_name="plate_log.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )